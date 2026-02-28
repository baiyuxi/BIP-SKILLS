#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模板写入模块
负责将数据写入模板并生成最终文件
"""

import os
import sys
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 添加父目录到路径以便导入config
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from config.settings import Config


class TemplateWriter:
    """模板写入器"""

    def __init__(self, template_structure: Dict[str, Any]):
        """
        初始化模板写入器
        
        Args:
            template_structure: 模板结构信息
        """
        self.structure = template_structure
        self.workbook = Workbook()
    
    def create_readme(self):
        """创建ReadMe工作表"""
        if Config.WORKSHEETS['readme'] in self.workbook.sheetnames:
            ws = self.workbook[Config.WORKSHEETS['readme']]
        else:
            ws = self.workbook.create_sheet(Config.WORKSHEETS['readme'], 0)
        
        readme_content = """Excel模板使用指南:
模板格式:请严格遵守表格中既定字段顺序,不得擅自调整或增减表头字段,以确保模板合法性。
数据类型:单元格内请输入文本格式的数据,以确保数值不会因excel软件转换丢失精度。
必填标识:带星号(*)的字段为必须填写项,请确保不遗漏。
引用档案:黄色背景字段为引用档案(参照),请填写参照名称或编码。
日期格式:建议使用"YYYY-MM-DD"格式(例如:2021-06-06)来输入日期,确保日期信息的标准化。
特殊字段说明:浅绿色背景字段用于确保数据的唯一性。若无特定业务唯一性要求,则默认使用系统生成的ID(手工码)。若子孙表中有浅蓝色背景的手工码,它是用于关联主表与子表数据的关联标识。
导入过滤:"✓"标记用于指示该行数据在导入过程中将被忽略。此标识常见于业务说明、属性说明或错误明细文件中,以避免重复或错误数据的导入。
国际化格式支持:下载模板时,若启用了"我的首选"功能,下载的模板将自动应用"首选项"的时区、数值、时间等格式设置。在导入导入入时,系统将依据这些设置进行自动转换与存储。
总结:请遵循上述指南,以确保数据导入的顺利进行和结果的准确性"""
        
        # 设置单元格格式
        ws['A1'].value = readme_content
        ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['A1'].font = Font(name='宋体', size=10)
        ws.row_dimensions[1].height = 200  # 设置行高
        
        # 设置列宽
        ws.column_dimensions['A'].width = 100
    
    def write_duty_data(self, data: List[List[Any]]):
        """
        写入职务数据
        
        Args:
            data: 职务数据列表
        """
        if Config.WORKSHEETS['duty'] not in self.workbook.sheetnames:
            self.workbook.create_sheet(Config.WORKSHEETS['duty'])
        
        ws = self.workbook[Config.WORKSHEETS['duty']]
        
        # 设置列宽
        column_widths = {
            'A': 15, 'B': 10, 'C': 20, 'D': 20, 'E': 30,
            'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 50, 'K': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 写入数据行(从第9行开始)
        start_row = Config.DATA_START_ROW
        for row_idx, record in enumerate(data, start_row):
            for col_idx, value in enumerate(record):
                ws.cell(row=row_idx, column=col_idx + 1, value=value)
    
    def write_ref_data(self, ref_data: List[List[Any]]):
        """
        写入参照数据
        
        Args:
            ref_data: 参照数据列表
        """
        if Config.WORKSHEETS['ref'] not in self.workbook.sheetnames:
            self.workbook.create_sheet(Config.WORKSHEETS['ref'])
        
        ws = self.workbook[Config.WORKSHEETS['ref']]
        
        # 清空现有数据
        ws.delete_rows(1, ws.max_row)
        
        # 写入新数据
        for row_idx, record in enumerate(ref_data, start=1):
            for col_idx, value in enumerate(record):
                ws.cell(row=row_idx, column=col_idx + 1, value=value)
    
    def copy_template_structure(self, template_reader):
        """
        从模板读取器复制结构
        
        Args:
            template_reader: 模板读取器实例
        """
        template_reader.copy_all_sheets(self.workbook)
        
        # 移除默认的Sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
    
    def update_duty_data(self, data: List[List[Any]]):
        """
        更新职务数据(覆盖模板中的数据行)
        
        Args:
            data: 职务数据列表
        """
        if Config.WORKSHEETS['duty'] not in self.workbook.sheetnames:
            raise ValueError(f"工作表不存在: {Config.WORKSHEETS['duty']}")
        
        ws = self.workbook[Config.WORKSHEETS['duty']]
        
        # 清空第9行之后的数据
        if ws.max_row >= Config.DATA_START_ROW:
            ws.delete_rows(Config.DATA_START_ROW, ws.max_row - Config.DATA_START_ROW + 1)
        
        # 写入新数据
        self.write_duty_data(data)
    
    def save(self, output_path: Optional[str] = None) -> str:
        """
        保存工作簿
        
        Args:
            output_path: 输出文件路径,如果为None则使用默认路径
            
        Returns:
            str: 保存的文件路径
        """
        if output_path is None:
            output_path = Config.get_output_path()
        
        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        self.workbook.save(output_path)
        return output_path
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()
