#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook("assets/模板_职务.xlsx")

# 检查所有工作表
print("工作表列表:", wb.sheetnames)
print()

# 检查职务工作表
if "职务" in wb.sheetnames:
    ws = wb["职务"]
    print("=== 职务工作表 ===")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
    print()

    # 打印前10行，寻找表头和 * 标记
    for row in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(15, ws.max_column + 1)):
            cell_value = ws.cell(row, col).value
            row_data.append(str(cell_value) if cell_value else "")
        print(f"行{row:2d}: {row_data}")

print()

# 检查 ReadMe 工作表
if "ReadMe" in wb.sheetnames:
    ws = wb["ReadMe"]
    print("=== ReadMe 工作表 ===")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
    print()

    # 打印所有内容
    for row in range(1, ws.max_row + 1):
        row_data = []
        for col in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row, col).value
            row_data.append(str(cell_value) if cell_value else "")
        if any(row_data):  # 只打印非空行
            print(f"行{row:2d}: {row_data}")
