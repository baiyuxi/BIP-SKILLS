#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用友BIP职务档案导入模板生成器 - 简化版
"""

import os
import sys
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 模板路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(SCRIPT_DIR, "..", "assets", "模板_职务.xlsx")

# 职务类型前缀映射
JOB_TYPE_PREFIX = {
    "产品经理": "PM",
    "研发经理": "RDM", 
    "需求分析": "REQ",
    "软件测试": "TEST",
    "前端开发": "FE",
    "后端开发": "BE",
    "UI设计": "UI",
    "运维": "OPS",
    "产品": "PM",
    "研发": "RD",
    "开发": "DEV",
    "测试": "TEST",
    "设计": "DES",
}


def parse_duties(text: str) -> list:
    """从文本中解析职务名称"""
    # 提取"职务:xxx"或"包括xxx"格式的内容
    pattern = r'职务[：:]([^，。,\n]+)|包括[：:]?\s*([^，。,\n]+)'
    matches = re.findall(pattern, text)
    
    duties = []
    for match in matches:
        duty = match[0] or match[1]
        # 处理顿号分隔的多个职务
        for d in duty.split('、'):
            d = d.strip()
            if d:
                duties.append(d)
    return duties


def generate_code(duty_name: str, index: int) -> str:
    """生成职务编码"""
    # 查找匹配的前缀
    prefix = "JOB"
    for key, value in JOB_TYPE_PREFIX.items():
        if key in duty_name:
            prefix = value
            break
    
    # 生成编码: 前缀 + 月日时分秒 + 序号
    now = datetime.now().strftime("%m%d%H%M%S")
    return f"{prefix}{now}{index:03d}"


def generate_duties_desc(duty_name: str) -> str:
    """生成职务描述"""
    desc_map = {
        "产品经理": "负责产品规划与设计,制定产品路线图,协调研发团队",
        "研发经理": "负责研发团队管理,推动技术创新,确保项目按时交付",
        "需求分析": "负责收集与分析用户需求,编写需求文档,协调各方资源",
        "软件测试": "负责软件测试工作,制定测试计划,保障产品质量",
        "前端开发": "负责前端开发工作,实现用户界面,优化用户体验",
        "后端开发": "负责后端开发工作,构建系统架构,保证服务稳定",
        "UI设计": "负责UI设计工作,制定设计规范,提升产品美观度",
    }
    
    for key, desc in desc_map.items():
        if key in duty_name:
            return desc[:50]  # 限制50字
    
    return f"负责{duty_name}相关工作"


def load_template() -> load_workbook:
    """加载模板"""
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_FILE}")
    return load_workbook(TEMPLATE_FILE)


def get_organizations(wb) -> list:
    """获取可选组织列表"""
    if "参照" in wb.sheetnames:
        ws = wb["参照"]
        orgs = []
        for row in range(2, ws.max_row + 1):
            org_name = ws.cell(row, 1).value
            if org_name:
                orgs.append(org_name)
        return orgs
    return []


def check_required_fields(wb) -> dict:
    """
    检查模板中的必填字段
    
    Returns:
        字段名到列号的映射字典
    """
    required_fields = {}
    
    if "职务" in wb.sheetnames:
        ws = wb["职务"]
        
        # 从第1行开始查找表头（通常在第8行左右）
        for row in range(1, 10):
            # 查找包含 "*" 标记的列，这些是必填字段
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    # 检查是否包含 "*" 标记
                    if "*" in cell_value:
                        # 提取字段名（去掉 * 标记）
                        field_name = cell_value.replace("*", "").strip()
                        required_fields[field_name] = col
        
        # 从 ReadMe 工作表获取必填字段信息
        if "ReadMe" in wb.sheetnames:
            readme_ws = wb["ReadMe"]
            for row in range(1, readme_ws.max_row + 1):
                cell = readme_ws.cell(row, 1)
                if cell.value and "必填" in str(cell.value):
                    # 解析必填字段说明
                    break
    
    return required_fields


def generate(user_input: str, org_name: str = None, remarks: dict = None, job_types: dict = None, duties_desc: dict = None) -> str:
    """
    生成职务导入模板

    Args:
        user_input: 用户输入，如"我需要导入职务:产品经理、研发经理"
        org_name: 所属组织名称，如"企业账号级##global00"
        remarks: 备注字典，格式: {"职务名": "备注内容"}
        job_types: 职务类别字典，格式: {"职务名": "职务类别"}
        duties_desc: 职责描述字典，格式: {"职务名": "职责描述"}

    Returns:
        生成的文件路径
    """
    # 解析职务列表
    duties = parse_duties(user_input)
    if not duties:
        raise ValueError("未能从输入中解析到职务信息")

    # 加载模板
    wb = load_template()

    # 检查必填字段
    required_fields = check_required_fields(wb)
    if required_fields:
        print(f"✓ 检测到必填字段: {list(required_fields.keys())}")

    # 如果没有指定组织，获取第一个可用组织
    if not org_name:
        orgs = get_organizations(wb)
        if orgs:
            org_name = orgs[0]
        else:
            raise ValueError("未找到可选组织，请手动指定")

    # 验证所有必填字段
    missing_fields = []
    if "备注" in required_fields and not remarks:
        missing_fields.append("备注")
    if "职务类别" in required_fields and not job_types:
        missing_fields.append("职务类别")
    if "职责" in required_fields and not duties_desc:
        missing_fields.append("职责")

    if missing_fields:
        raise ValueError(f"【严重错误】以下必填字段未提供: {', '.join(missing_fields)}\n"
                        f"请使用完整参数调用脚本，例如：\n"
                        f"python scripts/generate.py \"职务列表\" \"组织\" "
                        f"'{{\"职务名\": \"备注\"}}' '{{\"职务名\": \"职务类别\"}}' '{{\"职务名\": \"职责\"}}'")

    # 获取或创建职务工作表
    if "职务" in wb.sheetnames:
        ws = wb["职务"]
    else:
        ws = wb.create_sheet("职务")
    
    # 从第9行开始写入数据
    start_row = 9
    
    # 清空已有数据(保留表头)
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, 12):
            ws.cell(row, col).value = None
    
    # 写入新数据
    for i, duty_name in enumerate(duties):
        row = start_row + i
        
        # 生成职务编码
        code = generate_code(duty_name, i + 1)
        
        # 生成职务描述
        duties_desc = generate_duties_desc(duty_name)
        
        # 写入数据: ERROR_MSG, filter_flag, code, name, org_id_name, ...
        ws.cell(row, 1).value = None  # ERROR_MSG
        ws.cell(row, 2).value = None  # filter_flag (不设置'√')
        ws.cell(row, 3).value = code  # 职务编号
        ws.cell(row, 4).value = duty_name  # 职务名称
        ws.cell(row, 5).value = org_name  # 所属组织

        # 写入职务类别（如果必填）
        if "职务类别" in required_fields and job_types and duty_name in job_types:
            ws.cell(row, 6).value = job_types[duty_name]

        # 写入职责描述（如果必填）
        if "职责" in required_fields and duties_desc and duty_name in duties_desc:
            ws.cell(row, 10).value = duties_desc[duty_name]

        # 写入备注（如果必填）
        if "备注" in required_fields and remarks and duty_name in remarks:
            ws.cell(row, 11).value = remarks[duty_name]
    
    # 生成输出文件名
    timestamp = datetime.now().strftime("%m%d%H%M%S")
    output_file = f"模板_职务_{timestamp}.xlsx"
    
    # 保存文件
    wb.save(output_file)
    print(f"✓ 模板已生成: {output_file}")
    print(f"  - 职务数量: {len(duties)}")
    print(f"  - 所属组织: {org_name}")
    
    return output_file


if __name__ == "__main__":
    # 测试
    if len(sys.argv) > 1:
        user_input = sys.argv[1]
        org = sys.argv[2] if len(sys.argv) > 2 else None

        # 解析可选参数（JSON格式）
        import json
        remarks = None
        job_types = None
        duties_desc = None

        if len(sys.argv) > 3:
            try:
                remarks = json.loads(sys.argv[3])
            except json.JSONDecodeError:
                pass

        if len(sys.argv) > 4:
            try:
                job_types = json.loads(sys.argv[4])
            except json.JSONDecodeError:
                pass

        if len(sys.argv) > 5:
            try:
                duties_desc = json.loads(sys.argv[5])
            except json.JSONDecodeError:
                pass

        generate(user_input, org, remarks, job_types, duties_desc)
    else:
        # 默认测试
        print("测试1: 检查必填字段")
        try:
            generate("我需要导入职务:产品经理")
        except ValueError as e:
            print(f"✗ {e}")

        print("\n测试2: 完整参数")
        remarks = {"产品经理": "负责产品规划"}
        job_types = {"产品经理": "管理岗"}
        duties_desc = {"产品经理": "制定产品路线图"}
        generate("我需要导入职务:产品经理", None, remarks, job_types, duties_desc)
