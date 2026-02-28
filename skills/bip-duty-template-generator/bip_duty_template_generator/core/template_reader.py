#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模板读取与解析模块
负责从参考模板中读取结构和数据
"""

import os
import sys
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook

# 添加父目录到路径以便导入config
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from config.settings import Config


class TemplateReader:
    """模板读取器"""

    def __init__(self, template_path: str = None):
        """
        初始化模板读取器
        
        Args:
            template_path: 模板文件路径,如果为None则使用默认模板路径
        """
        self.template_path = template_path or Config.TEMPLATE_PATH
        
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"模板文件不存在: {self.template_path}")
        
        self.workbook = None
        self.structure = None
        
    def load(self):
        """
        加载模板文件

        每次调用都会重新读取最新的模板文件,确保模板更新后能立即生效
        """
        # 检查文件是否存在
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"模板文件不存在: {self.template_path}")

        # 每次都重新加载最新的模板文件
        self.workbook = load_workbook(self.template_path, data_only=True)

        # 清除缓存的结构信息
        self.structure = None

        return self
    
    def close(self):
        """关闭模板文件"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def get_structure(self) -> Dict[str, Any]:
        """
        获取模板结构信息
        
        Returns:
            Dict[str, Any]: 模板结构
        """
        if not self.workbook:
            self.load()
        
        structure = {
            'worksheets': {},
            'field_info': {}
        }
        
        # 解析职务工作表
        if Config.WORKSHEETS['duty'] in self.workbook.sheetnames:
            duty_sheet = self.workbook[Config.WORKSHEETS['duty']]
            structure['worksheets']['duty'] = self._parse_duty_sheet(duty_sheet)
        
        # 解析参照工作表
        if Config.WORKSHEETS['ref'] in self.workbook.sheetnames:
            ref_sheet = self.workbook[Config.WORKSHEETS['ref']]
            structure['worksheets']['ref'] = self._parse_ref_sheet(ref_sheet)
        
        # 解析其他工作表
        for sheet_name in self.workbook.sheetnames:
            if sheet_name not in structure['worksheets']:
                sheet = self.workbook[sheet_name]
                structure['worksheets'][sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column
                }
        
        self.structure = structure
        return structure
    
    def _parse_duty_sheet(self, sheet) -> Dict[str, Any]:
        """解析职务工作表"""
        sheet_info = {
            'headers': {},
            'required_fields': [],
            'field_names': [],
            'total_columns': sheet.max_column
        }
        
        # 读取字段信息行(第4行:字段编码, 第6行:字段名称, 第8行:数据类型)
        field_codes = []
        field_names = []
        required_fields = []
        
        # 第4行:字段编码
        for cell in sheet[4]:
            if cell.value:
                field_codes.append(str(cell.value))
        
        # 第6行:字段名称(带*标记必填)
        for idx, cell in enumerate(sheet[6]):
            if cell.value:
                field_name = str(cell.value)
                field_names.append(field_name)
                
                # 检查是否为必填字段
                if '*' in field_name:
                    clean_name = field_name.replace('*', '').strip()
                    required_fields.append({
                        'name': clean_name,
                        'code': field_codes[idx] if idx < len(field_codes) else '',
                        'column': idx
                    })
        
        sheet_info['headers']['field_codes'] = field_codes
        sheet_info['headers']['field_names'] = field_names
        sheet_info['required_fields'] = required_fields
        sheet_info['field_names'] = field_names
        
        return sheet_info
    
    def _parse_ref_sheet(self, sheet) -> Dict[str, Any]:
        """解析参照工作表"""
        ref_info = {
            'organizations': [],
            'data': []
        }

        # 从第3行开始读取数据(跳过表头:第1行字段说明,第2行列名)
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row and row[0]:  # 第1列有值
                ref_info['data'].append(row)

                # 尝试提取组织信息
                if len(row) >= 2:
                    org_code = str(row[0]).strip()
                    org_name = str(row[1]).strip() if row[1] else ''
                    if org_name and org_name != '名称':  # 过滤表头
                        ref_info['organizations'].append(f"{org_name}##{org_code}")

        return ref_info
    
    def get_organizations(self) -> List[str]:
        """
        获取可用的组织列表
        
        Returns:
            List[str]: 组织列表
        """
        structure = self.get_structure()
        return structure.get('worksheets', {}).get('ref', {}).get('organizations', [])
    
    def get_required_fields(self) -> List[Dict[str, str]]:
        """
        获取必填字段列表(从模板的*标记中读取)

        Returns:
            List[Dict[str, str]]: 必填字段列表,每个元素包含name和code
        """
        structure = self.get_structure()
        return structure.get('worksheets', {}).get('duty', {}).get('required_fields', [])
    
    def get_field_mapping(self) -> Dict[str, int]:
        """
        获取字段编码到列索引的映射
        
        Returns:
            Dict[str, int]: 字段编码到列索引的映射
        """
        structure = self.get_structure()
        field_codes = structure.get('worksheets', {}).get('duty', {}).get('headers', {}).get('field_codes', [])
        return {code: idx for idx, code in enumerate(field_codes)}
    
    def copy_sheet_to_workbook(self, sheet_name: str, target_workbook, target_sheet_name: str = None):
        """
        复制工作表到目标工作簿
        
        Args:
            sheet_name: 源工作表名称
            target_workbook: 目标工作簿
            target_sheet_name: 目标工作表名称(如果为None则使用源名称)
        """
        if not self.workbook:
            self.load()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        
        src_sheet = self.workbook[sheet_name]
        target_name = target_sheet_name or sheet_name
        
        # 创建或获取目标工作表
        if target_name in target_workbook.sheetnames:
            dst_sheet = target_workbook[target_name]
        else:
            dst_sheet = target_workbook.create_sheet(target_name)
        
        # 复制内容和样式
        for row in src_sheet.iter_rows():
            for cell in row:
                dst_cell = dst_sheet.cell(row=cell.row, column=cell.column)
                dst_cell.value = cell.value
                
                if cell.has_style:
                    dst_cell.font = cell.font.copy()
                    dst_cell.border = cell.border.copy()
                    dst_cell.fill = cell.fill.copy()
                    dst_cell.number_format = cell.number_format.copy()
                    dst_cell.protection = cell.protection.copy()
                    dst_cell.alignment = cell.alignment.copy()
    
    def copy_all_sheets(self, target_workbook):
        """
        复制所有工作表到目标工作簿
        
        Args:
            target_workbook: 目标工作簿
        """
        if not self.workbook:
            self.load()
        
        for sheet_name in self.workbook.sheetnames:
            self.copy_sheet_to_workbook(sheet_name, target_workbook)
    
    def __enter__(self):
        """上下文管理器入口"""
        self.load()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()
