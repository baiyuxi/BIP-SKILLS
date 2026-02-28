#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成自定义职务档案导入模板
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os

# 默认所属组织
DEFAULT_ORGANIZATION = "企业账号级##global00"

# 用户提供的职务数据
user_duties = [
    {
        'code': datetime.now().strftime('%Y%m%d%H%M%S'),
        'name': '产品总监',
        'org': DEFAULT_ORGANIZATION,
        'description': '负责公司整体产品战略规划与管理，带领团队实现产品目标'
    },
    {
        'code': datetime.now().strftime('%Y%m%d%H%M%S'),
        'name': '总应用架构师',
        'org': DEFAULT_ORGANIZATION,
        'description': '负责应用系统整体架构设计与技术选型，确保架构合理性'
    },
    {
        'code': datetime.now().strftime('%Y%m%d%H%M%S'),
        'name': '应用架构师',
        'org': DEFAULT_ORGANIZATION,
        'description': '负责具体应用模块的架构设计与技术实现指导'
    },
    {
        'code': datetime.now().strftime('%Y%m%d%H%M%S'),
        'name': '需求分析师',
        'org': DEFAULT_ORGANIZATION,
        'description': '负责业务需求分析与建模，编写需求文档并推动实现'
    }
]

def generate_custom_template():
    """生成自定义职务模板"""
    
    template_path = '/Users/baiyuxi/Documents/YONYOUWORK/AI_BIP_PM/.trae/skills/bip-duty-template-generator/references/模板_职务.xlsx'
    base_dir = '/Users/baiyuxi/Documents/YONYOUWORK/AI_BIP_PM/产品规划过程/101项目交付过程/解决方案/标准Excel导入模板'
    
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)
    
    timestamp = datetime.now().strftime('%m%d')
    output_path = os.path.join(base_dir, f'模板_职务_{timestamp}.xlsx')
    
    # 读取参考模板
    ref_wb = load_workbook(template_path)
    
    # 创建新工作簿
    from openpyxl import Workbook
    wb = Workbook()
    
    # 复制所有sheet
    for sheet_idx, sheet_name in enumerate(ref_wb.sheetnames):
        if sheet_idx == 0:
            src_ws = ref_wb[sheet_name]
            dst_ws = wb.active
            dst_ws.title = sheet_name
            
            for row_idx, row in enumerate(src_ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell is not None:
                        dst_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                        if cell.has_style:
                            dst_ws.cell(row=row_idx, column=col_idx).font = cell.font
                            dst_ws.cell(row=row_idx, column=col_idx).border = cell.border
                            dst_ws.cell(row=row_idx, column=col_idx).fill = cell.fill
                            dst_ws.cell(row=row_idx, column=col_idx).number_format = cell.number_format
                            dst_ws.cell(row=row_idx, column=col_idx).protection = cell.protection
                            dst_ws.cell(row=row_idx, column=col_idx).alignment = cell.alignment
        else:
            src_ws = ref_wb[sheet_name]
            dst_ws = wb.create_sheet(title=sheet_name)
            
            for row_idx, row in enumerate(src_ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell is not None:
                        dst_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                        if cell.has_style:
                            dst_ws.cell(row=row_idx, column=col_idx).font = cell.font
                            dst_ws.cell(row=row_idx, column=col_idx).border = cell.border
                            dst_ws.cell(row=row_idx, column=col_idx).fill = cell.fill
                            dst_ws.cell(row=row_idx, column=col_idx).number_format = cell.number_format
                            dst_ws.cell(row=row_idx, column=col_idx).protection = cell.protection
                            dst_ws.cell(row=row_idx, column=col_idx).alignment = cell.alignment
    
    ref_wb.close()
    
    # 更新职务工作表
    if '职务' in wb.sheetnames:
        duty_ws = wb['职务']
        
        # 设置列宽
        column_widths = {
            'A': 15, 'B': 10, 'C': 15, 'D': 15, 'E': 15, 
            'F': 15, 'G': 15, 'H': 30, 'I': 30, 'J': 30, 'K': 30
        }
        for col, width in column_widths.items():
            duty_ws.column_dimensions[col].width = width
        
        # 填充用户数据
        start_row = 9
        for duty in user_duties:
            row_data = [
                '√',  # filter_flag
                None,  # bd.duty.Duty
                duty['code'],  # code
                duty['name'],  # name
                duty['org'],  # org_id_name
                None,  # jobtype_id_name
                None,  # jobgrade_id_name
                None,  # maxrank_id_name
                None,  # minrank_id_name
                duty['description'],  # duties
                None  # memo
            ]
            for col_idx, value in enumerate(row_data, start=1):
                if col_idx <= 11:
                    duty_ws.cell(row=start_row, column=col_idx, value=value)
            start_row += 1
        
        # 清空剩余行
        for row_idx in range(start_row, 12):
            for col_idx in range(1, 11):
                duty_ws.cell(row=row_idx, column=col_idx, value=None)
    
    wb.save(output_path)
    
    return output_path

if __name__ == '__main__':
    result_path = generate_custom_template()
    print(f"文件已生成: {result_path}")
    
    # 打印生成的数据
    print("\n生成的职务数据:")
    for i, duty in enumerate(user_duties, 1):
        print(f"{i}. {duty['name']} - 编码: {duty['code']}")
        print(f"   描述: {duty['description']}")
