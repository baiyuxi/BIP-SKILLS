#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成职务档案导入模板 - 自定义版本
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# 默认所属组织
DEFAULT_ORGANIZATION = "企业账号级##global00"


def generate_duty_template():
    """生成职务档案导入模板"""
    
    base_dir = '/Users/baiyuxi/Documents/YONYOUWORK/AI_BIP_PM/产品规划过程/101项目交付过程/解决方案/标准Excel导入模板'
    
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)
    
    timestamp = datetime.now().strftime('%m%d')
    output_path = os.path.join(base_dir, f'模板_职务_{timestamp}.xlsx')
    
    ref_template_path = '/Users/baiyuxi/Documents/YONYOUWORK/AI_BIP_PM/.trae/skills/bip-duty-template-generator/references/模板_职务.xlsx'
    
    ref_wb = load_workbook(ref_template_path)
    
    wb = Workbook()
    
    for sheet_idx, sheet_name in enumerate(ref_wb.sheetnames):
        if sheet_idx == 0:
            src_ws = ref_wb[sheet_name]
            dst_ws = wb.active
            dst_ws.title = sheet_name
            copy_sheet(src_ws, dst_ws)
        else:
            src_ws = ref_wb[sheet_name]
            dst_ws = wb.create_sheet(title=sheet_name)
            copy_sheet(src_ws, dst_ws)
    
    ref_wb.close()
    
    user_data = [
        ['√', '', '001', '初级需求分析', DEFAULT_ORGANIZATION, '', '', '', '', '', '负责收集和记录业务部门的基本需求，协助中级和高级需求分析师进行需求分析工作。'],
        ['√', '', '002', '中级需求分析', DEFAULT_ORGANIZATION, '', '', '', '', '', '负责中等复杂度业务模块的需求分析，进行需求调研、分析和文档编写，协调各方达成需求共识。'],
        ['√', '', '003', '高级需求分析', DEFAULT_ORGANIZATION, '', '', '', '', '', '负责复杂业务系统的需求分析，主导需求调研和分析，制定需求方案，指导初级和中级分析师工作。'],
    ]
    
    if '职务' in wb.sheetnames:
        duty_ws = wb['职务']
        
        column_widths = {
            'A': 15, 'B': 10, 'C': 15, 'D': 15, 'E': 15, 
            'F': 15, 'G': 15, 'H': 30, 'I': 30, 'J': 30, 'K': 30
        }
        for col, width in column_widths.items():
            duty_ws.column_dimensions[col].width = width
        
        start_row = 9
        for row_data in user_data:
            for col_idx, value in enumerate(row_data, start=1):
                if col_idx <= 11:
                    duty_ws.cell(row=start_row, column=col_idx, value=value)
            start_row += 1
        
        for row_idx in range(start_row, 12):
            for col_idx in range(1, 12):
                duty_ws.cell(row=row_idx, column=col_idx, value=None)
    
    wb.save(output_path)
    
    return output_path


def copy_sheet(src_ws, dst_ws):
    """复制工作表内容和样式"""
    for row_idx, row in enumerate(src_ws.iter_rows(values_only=False), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell is not None:
                dst_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                if cell.has_style:
                    dst_ws.cell(row=row_idx, column=col_idx).font = cell.font
                    dst_ws.cell(row=row_idx, column=col_idx).border = cell.border
                    dst_ws.cell(row=row_idx, column=col_idx).fill = cell.fill
                    dst_ws.cell(row=row_idx, column=col_idx).number_format = cell.number_format
                    dst_ws.cell(row=row_idx, column=col_idx).protection = cell.protection
                    dst_ws.cell(row=row_idx, column=col_idx).alignment = cell.alignment


if __name__ == '__main__':
    result_path = generate_duty_template()
    print(f"文件已生成: {result_path}")
